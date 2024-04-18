import os
import re
import sys
import time
import json
import math
import logging
import sqlite3
import requests
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook, styles
from openpyxl.utils import get_column_letter

class Config:
    def __init__(self) -> None:
        '''配置类'''
        self.config_file = f'./config.json'
        self.default_config_dict = {
            'host': '127.0.0.1',
            'port': 8840,
            'database_path': './data.db',
            'main_token': '',
            'auth_token': '',
            'output_file': './小班数据.xlsx',
            'daily_record': '59 23 * * *',
        }
        self.initConfig()
        self.raw = self.read()
        self.host = self.raw.get('host', '')
        self.port = self.raw.get('port', '')
        self.database_path = self.raw.get('database_path', '')
        self.main_token = self.raw.get('main_token', '')
        self.auth_token = self.raw.get('auth_token', '')
        self.output_file = self.raw.get('output_file', '')
        self.daily_record = self.raw.get('daily_record', '')
        self.verify()

    def initConfig(self):
        '''初始化配置文件'''
        try:
            if path := os.path.dirname(self.config_file):
                os.makedirs(path, exist_ok=True)
            open(self.config_file, encoding='utf-8')
        except:
            json.dump(self.default_config_dict, open(self.config_file, mode='w', encoding='utf-8'), ensure_ascii=False, indent=2)
            logging.info('初次启动，已在当前执行目录生成配置文件，请修改配置后再次启动，程序会在5秒后自动退出')
            time.sleep(5)
            sys.exit(0)

    def read(self, key: str=None) -> list | dict | str | int | bool:
        '''获取指定配置'''
        try:
            json_data = json.load(open(self.config_file, encoding='utf-8'))
            if key:
                json_data = json.load(open(self.config_file, encoding='utf-8')).get(key)
            return json_data
        except Exception as e:
            logging.error(f'配置文件读取异常: {e}，程序会在5秒后自动退出')
            time.sleep(5)
            sys.exit(0)

    def save(self, key: str, value: list | dict | str | int | bool) -> None:
        '''保存指定配置文件'''
        try:
            json_data = json.load(open(self.config_file, encoding='utf-8'))
            json_data[key] = value
            json.dump(json_data, open(self.config_file, mode='w', encoding='utf-8'), ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f'保存配置文件发生错误\n {e}')

    def verify(self):
        '''验证配置文件的完整性'''
        value = None
        if self.host == '':
            key = 'host'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.host = value
        if self.port == '':
            key = 'port'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.port = value
        if self.database_path == '':
            key = 'database_path'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.database_path = value
        if self.main_token == '':
            key = 'main_token'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.main_token = value
        if self.auth_token == '':
            key = 'auth_token'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.auth_token = value
        if self.output_file == '':
            key = 'output_file'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.output_file = value
        if self.daily_record == '':
            key = 'daily_record'
            value = self.default_config_dict[key]
            self.save(key, value)
            self.daily_record = value

    def getInfo(self) -> dict:
        '''获取配置文件相关状态信息'''
        return {
            'daily_record': self.daily_record
        }

class BCZ:
    def __init__(self, config: Config) -> None:
        '''小班解析类'''
        self.main_token = config.main_token
        self.auth_token = config.auth_token
        self.invalid_pattern = r'[\000-\010]|[\013-\014]|[\016-\037]'
        self.group_list_url = 'https://group.baicizhan.com/group/own_groups'
        self.group_detail_url = 'https://group.baicizhan.com/group/information'
        self.user_info_url = 'https://social.baicizhan.com/api/deskmate/personal_details'

    def getInfo(self) -> dict:
        '''获取运行信息'''
        main_info = self.getOwnInfo(self.main_token)
        token_valid = False
        if main_info['uid']:
            token_valid = True
        return {
            'token_valid': token_valid,
            'uid': main_info['uid'],
            'name': main_info['name'],
        }

    def getOwnInfo(self, token: str) -> dict:
        '''获取当前用户信息'''
        data = {
            'uid': None,
            'name': None,
        }
        url = 'https://social.baicizhan.com/api/deskmate/home_page'
        headers = {'Cookie': f'access_token="{token}"'}
        response = requests.get(url, headers=headers, timeout=5)
        try:
            response_json = response.json()
            if response_json.get('code') == 1:
                data['uid'] = response_json['data']['mine']['uniqueId']
                data['name'] = response_json['data']['mine']['name']
        except ValueError as e:
            logging.error(f'请求API[{url}]异常!\n{response.text}')
        return data
        
    def getUserInfo(self, user_id: str = None) -> dict | None:
        '''获取指定用户信息'''
        if not self.user_id and user_id:
            return
        elif not user_id:
            user_id = self.user_id
        url = f'{self.user_info_url}?uniqueId={user_id}'
        headers = {'Cookie': f'access_token="{self.main_token}"'}
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code != 200 or response.json().get('code') != 1:
            msg = f'获取我的小班信息失败! 用户不存在或主授权令牌无效'
            logging.error(f'{msg}\n{response.text}')
            raise Exception(msg)
        user_info = response.json()['data']
        return user_info

    def getUserGroupInfo(self, user_id: str = None) -> dict | None:
        '''获取我的小班信息'''
        if not self.user_id and user_id:
            return
        elif not user_id:
            user_id = self.user_id
        url = f'{self.group_list_url}?uniqueId={user_id}'
        headers = {'Cookie': f'access_token="{self.main_token}"'}
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code != 200 or response.json().get('code') != 1:
            msg = f'获取我的小班信息失败! 用户不存在或主授权令牌无效'
            logging.error(f'{msg}\n{response.text}')
            raise Exception(msg)
        group_info = response.json()['data']
        group_list = group_info.get('list') if group_info else []
        group_dict = {}
        self.data_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        for group in group_list:
            group_id = group['id']
            group_name = group['name'] if group['name'] else ''
            group_name = re.sub(self.invalid_pattern, '', group_name)
            introduction = group['introduction'] if group['introduction'] else ''
            introduction = re.sub(self.invalid_pattern, '', introduction)
            avatar_frame = group['avatarFrame']['frame'] if group.get('avatarFrame') else ''
            group_dict[group_id] = {
                'name': group_name,
                'shareKey': group['shareKey'],
                'introduction': introduction,
                'leader': '',
                'leaderId': '',
                'memberCount': group['memberCount'],
                'countLimit': group['countLimit'],
                'todayDakaCount': group['todayDakaCount'],
                'finishingRate': group['finishingRate'],
                'createdTime': group['createdTime'],
                'rank': group['rank'],
                'type': group['type'],
                'avatar': group['avatar'],
                'avatarFrame': avatar_frame,
                'dataTime': self.data_time,
            }
        return group_dict

    def getGroupInfo(self, share_key: str) -> dict | None:
        '''获取小班信息'''
        info = {}
        self.data_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        url = f'{self.group_detail_url}?shareKey={share_key}'
        headers = {'Cookie': f'access_token="{self.main_token}"'}
        main_response = requests.get(url, headers=headers, timeout=5)
        headers = {'Cookie': f'access_token="{self.auth_token}"'}
        auth_response = requests.get(url, headers=headers, timeout=5)
        if main_response.status_code != 200 or main_response.json().get('code') != 1:
            msg = f'获取分享码为{share_key}的小班信息失败! 小班不存在或主授权令牌无效'
            logging.error(f'{msg}\n{main_response.text}')
            raise Exception(msg)
        main_data = main_response.json()['data']

        group_info = main_data.get('groupInfo') if main_data else []
        group_id = group_info['id']
        group_name = re.sub(self.invalid_pattern, '', group_info['name']) if group_info['name'] else ''
        introduction = re.sub(self.invalid_pattern, '', group_info['introduction']) if group_info['introduction'] else ''
        notice = re.sub(self.invalid_pattern, '', group_info['notice']) if group_info['notice'] else ''
        avatar_frame = group_info['avatarFrame']['frame'] if group_info.get('avatarFrame') else ''
        info = {
            'id': group_id,
            'name': group_name,
            'shareKey': group_info['shareKey'],
            'introduction': introduction,
            'leader': '',
            'leaderId': '',
            'memberCount': group_info['memberCount'],
            'countLimit': group_info['countLimit'],
            'todayDakaCount': group_info['todayDakaCount'],
            'finishingRate': group_info['finishingRate'],
            'createdTime': group_info['createdTime'],
            'rank': group_info['rank'],
            'type': group_info['type'],
            'avatar': group_info['avatar'],
            'avatarFrame': avatar_frame,
            'notice': notice,
            'dataTime': self.data_time,
        }

        today_date = main_data.get('todayDate') if main_data else ''
        today_daka_count = 0
        main_member_list = main_data.get('members') if main_data else []
        member_dict = {}
        for member in main_member_list:
            member_id = member['uniqueId']
            nickname = re.sub(self.invalid_pattern, '', member['nickname'])
            completedTime = ''
            if member['completedTime']:
                today_daka_count += 1
                completedTime = time.strftime('%H:%M:%S', time.localtime(member['completedTime']))
            member_dict[member_id] = {
                'group_id': group_id,
                'group_name': group_name,
                'nickname': nickname,
                'group_nickname': '',
                'bookName': member['bookName'],
                'todayWordCount': member['todayWordCount'],
                'completedTimes': member['completedTimes'],
                'completedTime': completedTime,
                'durationDays': member['durationDays'],
                'todayStudyCheat': '是' if member['todayStudyCheat'] else '否',
                'todayDate': today_date,
                'dataTime': self.data_time,
            }
            if member['leader']:
                info['leader'] = member['nickname']
                info['leaderId'] = member_id

        auth_data = auth_response.json()['data']
        auth_member_list = auth_data.get('members') if auth_data else []
        for member in auth_member_list:
            member_id = member['uniqueId']
            nickname = re.sub(self.invalid_pattern, '', member['nickname'])
            if member_id in member_dict and member_dict[member_id]['nickname'] != nickname:
                member_dict[member_id]['group_nickname'] = member['nickname']
            else:
                member_dict[member_id]['group_nickname'] = ''
        if today_daka_count != 0:
            info['todayDakaCount'] = today_daka_count
        info['members'] = member_dict

        return info

    def updateGroupInfo(self, group_list: list[dict]) -> list:
        with ThreadPoolExecutor() as executor:
            share_keys = []
            for group_info in group_list:
                share_keys.append(group_info['shareKey'])
            results = executor.map(self.getGroupInfo, share_keys)
        for result in results:
            for group_info in group_list:
                if group_info['id'] == result['id']:
                     group_info.update(result)
        return group_list

    def getUserAllInfo(self, user_id: str = None) -> dict | None:
        '''获取指定用户所有信息'''
        user_info = self.getUserInfo(user_id)
        if not user_info:
            return
        user_group_dict = self.getUserGroupInfo(user_id)
        group_dict = {}
        for group_id, group in user_group_dict.items():
            group_dict[group_id] = self.getGroupInfo(group['shareKey'])
        user_info['group_dict'] = group_dict
        return user_info

class SQLite:
    def __init__(self, config: Config) -> None:
        '''数据库类'''
        self.db_path = config.database_path
        self.groups_table_sql = '''
        CREATE TABLE IF NOT EXISTS GROUPS (                   -- 小班表
            GROUP_ID INTEGER,                   -- 小班ID
            NAME TEXT,                          -- 小班名称
            SHARE_KEY TEXT,                     -- 小班分享码
            INTRO TEXT,                         -- 小班简介
            LEADER TEXT,                        -- 小班班长
            LEADER_ID TEXT,                     -- 班长ID
            MEMBER_COUNT INTEGER,               -- 当前人数
            COUNT_LIMIT INTEGER,                -- 人数上限
            TODAY_DAKA INTEGER,                 -- 今日打卡数
            FINISHING_RATE REAL,                -- 打卡率
            CREATED_TIME TEXT,                  -- 建立时间
            RANK INTEGER,                       -- 段位排行
            GROUP_TYPE INTEGER,                 -- 小班类型
            AVATAR TEXT,                        -- 小班头像
            AVATAR_FRAME TEXT,                  -- 小班像框
            DATA_TIME TEXT                      -- 采集时间
        );
        '''
        self.members_table_sql = '''
        CREATE TABLE IF NOT EXISTS MEMBERS (                   -- 成员表
            USER_ID INTEGER,                    -- 用户ID
            NICKNAME TEXT,                      -- 用户昵称
            GROUP_NICKNAME TEXT,                -- 班内昵称
            COMPLETED_TIME TEXT,                -- 打卡时间
            TODAY_DATE TEXT,                    -- 记录日期
            WORD_COUNT INTEGER,                 -- 今日词数
            STUDY_CHEAT INTEGER,                -- 是否作弊
            COMPLETED_TIMES INTEGER,            -- 打卡天数
            DURATION_DAYS INTEGER,              -- 入班天数
            BOOK_NAME TEXT,                     -- 学习词书
            GROUP_ID INTEGER,                   -- 小班ID
            GROUP_NAME TEXT,                    -- 小班昵称
            DATA_TIME TEXT                      -- 采集时间
        );
        '''
        self.observed_groups_table_sql = '''
        CREATE TABLE IF NOT EXISTS OBSERVED_GROUPS (                   -- 关注小班表
            GROUP_ID INTEGER,                   -- 小班ID
            NAME TEXT,                          -- 小班名称
            SHARE_KEY TEXT,                     -- 小班分享码
            INTRO TEXT,                         -- 小班简介
            LEADER TEXT,                        -- 小班班长
            LEADER_ID TEXT,                     -- 班长ID
            MEMBER_COUNT INTEGER,               -- 当前人数
            COUNT_LIMIT INTEGER,                -- 人数上限
            TODAY_DAKA INTEGER,                 -- 今日打卡数
            FINISHING_RATE REAL,                -- 打卡率
            CREATED_TIME TEXT,                  -- 建立时间
            RANK INTEGER,                       -- 段位排行
            GROUP_TYPE INTEGER,                 -- 小班类型
            AVATAR TEXT,                        -- 小班头像
            AVATAR_FRAME TEXT,                  -- 小班像框
            NOTICE TEXT,                        -- 小班公告
            DAILY_RECORD INTEGER,               -- 每日记录
            LATE_DAKA_TIME TEXT,                -- 晚打卡时间
            WEEKLY_ABSENCE INTEGER,             -- 每周缺卡上限
            AUTH_TOKEN TEXT                     -- 主授权令牌
        );
        '''
        self.init()

    def connect(self, db_path) -> sqlite3.Connection:
        '''连接数据库，并返回连接态'''
        try:
            if path := os.path.dirname(db_path):
                os.makedirs(path, exist_ok=True)
            conn = sqlite3.connect(db_path)
            conn.set_trace_callback(lambda statement: logging.debug(f'在{self.db_path}执行SQLite指令: {statement}'))
            return conn
        except sqlite3.Error:
            logging.error('数据库读取异常...无法正常运行，程序会在5秒后自动退出')
            time.sleep(5)
            sys.exit(0)

    def init(self) -> None:
        '''初始化不存在的库'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(self.groups_table_sql)
        cursor.execute(self.members_table_sql)
        cursor.execute(self.observed_groups_table_sql)
        conn.commit()

    def read(self, sql: str, param: list | tuple = ()) -> list:
        '''SQL执行读数据操作'''
        try:
            conn = self.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(sql, param)
            return cursor.fetchall()
        except sqlite3.DatabaseError as e:
            logging.error(f'读取数据库{self.db_path}出错: {e}')
            raise e

    def write(self, sql: str, param: list | tuple = ()) -> bool:
        '''SQL执行写数据操作'''
        try:
            conn = self.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(sql, param)
            conn.commit()
            return True
        except sqlite3.DatabaseError as e:
            logging.error(f'写入数据库{self.db_path}出错: {e}')
        return False

    def saveGroupInfo(self, group_list: list[dict]) -> None:
        '''仅保存小班详情'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        for group_info in group_list:
            cursor.execute(
                'INSERT OR IGNORE INTO GROUPS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (
                    group_info['id'],
                    group_info['name'],
                    group_info['shareKey'],
                    group_info['introduction'],
                    group_info['leader'],
                    group_info['leaderId'],
                    group_info['memberCount'],
                    group_info['countLimit'],
                    group_info['todayDakaCount'],
                    group_info['finishingRate'],
                    group_info['createdTime'],
                    group_info['rank'],
                    group_info['type'],
                    group_info['avatar'],
                    group_info['avatarFrame'],
                    group_info['dataTime'],
                )
            )
        conn.commit()
        self.saveMemberInfo(group_info['members'])

    def saveMemberInfo(self, member_dict: dict) -> None:
        '''仅保存成员详情'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        for id, member in member_dict.items():
            cursor.execute(
                'INSERT OR IGNORE INTO MEMBERS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (
                    id,
                    member['nickname'],
                    member['group_nickname'],
                    member['completedTime'],
                    member['todayDate'],
                    member['todayWordCount'],
                    member['todayStudyCheat'],
                    member['completedTimes'],
                    member['durationDays'],
                    member['bookName'],
                    member['group_id'],
                    member['group_name'],
                    member['dataTime'],
                )
            )
        conn.commit()

    def addObserveGroupInfo(self, group_list: list[dict]) -> None:
        '''增加关注小班信息'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        for group_info in group_list:
            cursor.execute(
                '''
                    INSERT INTO OBSERVED_GROUPS VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    group_info.get('id', 0),
                    group_info.get('name', ''),
                    group_info.get('shareKey', ''),
                    group_info.get('introduction', ''),
                    group_info.get('leader', ''),
                    group_info.get('leaderId', ''),
                    group_info.get('memberCount', 0),
                    group_info.get('countLimit', 0),
                    group_info.get('todayDakaCount', 0),
                    group_info.get('finishingRate', 0),
                    group_info.get('createdTime', ''),
                    group_info.get('rank', 1),
                    group_info.get('type', 0),
                    group_info.get('avatar', ''),
                    group_info.get('avatarFrame', ''),
                    group_info.get('notice', ''),
                    group_info.get('daily_record', 1),
                    group_info.get('late_daka_time', ''),
                    group_info.get('weekly_absence', 7),
                    group_info.get('auth_token', ''),
                )
            )
        conn.commit()

    def deleteObserveGroupInfo(self, group_id) -> None:
        '''删除关注小班信息'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            'DELETE FROM OBSERVED_GROUPS WHERE GROUP_ID = ?',
            (group_id)
        )
        conn.commit()

    def updateObserveGroupInfo(self, group_list: list[dict]) -> None:
        '''更新关注小班信息'''
        conn = self.connect(self.db_path)
        cursor = conn.cursor()
        for group_info in group_list:
            cursor.execute(
                '''
                    UPDATE OBSERVED_GROUPS SET
                        GROUP_ID = ?,
                        NAME = ?,
                        SHARE_KEY = ?,
                        INTRO = ?,
                        LEADER = ?,
                        LEADER_ID = ?,
                        MEMBER_COUNT = ?,
                        COUNT_LIMIT = ?,
                        TODAY_DAKA = ?,
                        FINISHING_RATE = ?,
                        CREATED_TIME = ?,
                        RANK = ?,
                        GROUP_TYPE = ?,
                        AVATAR = ?,
                        AVATAR_FRAME = ?,
                        NOTICE = ?
                    WHERE GROUP_ID = ?
                ''',
                (
                    group_info['id'],
                    group_info['name'],
                    group_info['shareKey'],
                    group_info['introduction'],
                    group_info['leader'],
                    group_info['leaderId'],
                    group_info['memberCount'],
                    group_info['countLimit'],
                    group_info['todayDakaCount'],
                    group_info['finishingRate'],
                    group_info['createdTime'],
                    group_info['rank'],
                    group_info['type'],
                    group_info['avatar'],
                    group_info['avatarFrame'],
                    group_info['notice'],
                    group_info['id'],
                )
            )
        conn.commit()

    def queryObserveGroupInfo(self, group_id: str = '', full_info: bool=False) -> dict:
        '''查询关注小班信息'''
        if group_id:
            result = self.read(
                f'SELECT * FROM OBSERVED_GROUPS WHERE GROUP_ID = ? ORDER BY GROUP_ID ASC',
                group_id,
            )
        else:
            result = self.read(f'SELECT * FROM OBSERVED_GROUPS ORDER BY GROUP_ID ASC')
        result_keys = [
            'id',
            'name',
            'shareKey',
            'introduction',
            'leader',
            'leaderId',
            'memberCount',
            'countLimit',
            'todayDakaCount',
            'finishingRate',
            'createdTime',
            'rank',
            'type',
            'avatar',
            'avatarFrame',
            'notice',
            'daily_record',
            'late_daka_time',
            'weekly_absence',
            'auth_token',
        ]
        group_info = []
        for item in result:
            group_info.append(dict(zip(result_keys, item)))
        if not full_info:
            for item in group_info:
                item['auth_token'] = len(item['auth_token']) * '*'
        return group_info

    def getDays(self) -> int:
        '''获取数据记录总天数'''
        result = self.read(
            f'SELECT COUNT(DISTINCT TODAY_DATE) FROM MEMBERS'
        )
        return result[0][0]

    def getInfo(self) -> dict:
        '''获取数据库相关状态信息'''
        return {
            'count': self.getMemberDataCount(),
            'groups': self.getDistinctGroupName(),
            'running_days': self.getDays()
        }

    def getGroupInfo(self) -> dict:
        '''获取记录小班详情'''
        result = self.read(
            f'''
                SELECT DISTINCT *
                FROM GROUPS A
                JOIN (
                    SELECT
                    GROUP_ID,
                    MAX(DATA_TIME) DATA_TIME 
                    FROM GROUPS 
                    GROUP BY GROUP_ID
                ) B ON A.GROUP_ID = B.GROUP_ID
                    AND A.DATA_TIME = B.DATA_TIME
                ORDER BY A.GROUP_ID ASC
            '''
        )
        result_keys = [
            'id',
            'name',
            'shareKey',
            'introduction',
            'leader',
            'leaderId',
            'memberCount',
            'countLimit',
            'todayDakaCount',
            'finishingRate',
            'createdTime',
            'rank',
            'type',
            'avatar',
            'avatarFrame'
        ]
        group_info = []
        for item in result:
            group_info.append(dict(zip(result_keys, item[:-1])))
        return group_info

    def getDistinctGroupName(self) -> list:
        return self.read(
            f'''
                SELECT DISTINCT
                    A.GROUP_ID,
                    A.GROUP_ID||'('||A.NAME||')' NAME
                FROM GROUPS A
                JOIN (
                    SELECT
                    GROUP_ID,
                    MAX(DATA_TIME) DATA_TIME 
                    FROM GROUPS 
                    GROUP BY GROUP_ID
                ) B ON A.GROUP_ID = B.GROUP_ID
                    AND A.DATA_TIME = B.DATA_TIME
                ORDER BY A.GROUP_ID ASC
            '''
        )

    def getSearchOption(self) -> dict:
        return {
            'groups': self.getDistinctGroupName()
        }

    def getMemberDataCount(self) -> int:
        return self.read(
            f'SELECT COUNT(*) FROM MEMBERS'
        )[0][0]

    def searchMemberTable(self, payload: dict, header: bool = False) -> list:
        '''查询用户信息表

        Args:
            payload (dict): {
                'page_num': 页数
                'page_count': 每页条数
                'group_id': 小班ID
                'group_name': 小班名称
                'sdate': 开始日期
                'edate': 结束日期
                'cheat': 是否作弊
                'completed': 是否完成
                'completed_time': 打卡时间
                'user_id': 用户ID
                'nickname': 用户昵称
            }

        Returns:
            list: 用户信息表
        '''        
        count_sql = 'SELECT COUNT(*) FROM MEMBERS WHERE 1=1'
        search_sql = 'SELECT * FROM MEMBERS WHERE 1=1'
        sql = ''
        param = []
        user_id = payload.get('user_id', '')
        nickname = payload.get('nickname', '')
        group_id = payload.get('group_id', '')
        group_name = payload.get('group_name', '')
        sdate = payload.get('sdate', '')
        edate = payload.get('edate', '')
        cheat = payload.get('cheat', '')
        completed = payload.get('completed', '')
        completed_time = payload.get('completed_time', '')
        if user_id != '':
            sql += ' AND USER_ID LIKE ?'
            param.append(f'%{user_id}%')
        if nickname != '':
            sql += ' AND (NICKNAME LIKE ? OR GROUP_NICKNAME LIKE ?)'
            param.append(f'%{nickname}%')
            param.append(f'%{nickname}%')
        if group_id != '':
            sql += ' AND GROUP_ID = ?'
            param.append(group_id)
        if group_name != '':
            sql += ' AND GROUP_NAME LIKE ?'
            param.append(f'%{group_name}%')
        if sdate != '' or edate != '':
            sdate = sdate if sdate else '0000-00-00'
            edate = edate if edate else '9999-12-31'
            sql += ' AND TODAY_DATE BETWEEN ? AND ?'
            param.append(sdate)
            param.append(edate)
        if cheat in ['true', 'false']:
            sql += ' AND STUDY_CHEAT = ?'
            param.append('是' if cheat == 'true' else '否')
        if completed in ['true', 'false']:
            if completed == 'true':
                sql += ' AND COMPLETED_TIME <> \'\''
            else:
                sql += ' AND COMPLETED_TIME = \'\''
        if completed_time != '':
            sql += ' AND (COMPLETED_TIME = \'\' OR COMPLETED_TIME > ?)'
            param.append(completed_time)
        count_sql += sql
        count_result = self.read(count_sql, param)
        count = count_result[0][0]
        page_max = 1
        page_num = 1
        page_count = 'unlimited'
        if payload.get('page_count', '') == '':
            pass
        elif payload.get('page_count', '') != 'unlimited':
            page_count = payload.get('page_count', 20)
            page_num = payload.get('page_num', 1)
            page_num = page_num if page_num else 1
            page_count = page_count if int(page_count) > 0 else 20
            page_max = math.ceil(int(count) / int(page_count))
            page_num = page_num if int(page_num) > 0 else 1
            page_num = page_num if int(page_num) < page_max else page_max
            sql += ' LIMIT ? OFFSET (? - 1) * ?'
            param.append(page_count)
            param.append(page_num)
            param.append(page_count)
        search_sql += sql
        result = self.read(search_sql, param)
        if header:
            result = [[
                '用户ID',
                '用户昵称',
                '班内昵称',
                '打卡时间',
                '记录日期',
                '今日词数',
                '是否作弊',
                '打卡天数',
                '入班天数',
                '学习词书',
                '小班ID',
                '小班名称',
                '采集时间',
            ]] + result
        return [result, count, page_max, page_num, page_count]


class Xlsx:
    def __init__(self, config: Config) -> None:
        '''表格数据操作类'''
        self.file_path = config.output_file
        if path := os.path.dirname(self.file_path):
            os.makedirs(path, exist_ok=True)

    def write(self, sheet_name: str, data: list, overwrite: bool = True) -> bool:
        '''数据写入到表格'''  
        if overwrite:
            self.wb = Workbook()
            self.wb.remove(self.wb['Sheet'])
        else:
            try:
                self.wb = load_workbook(self.file_path)
            except FileNotFoundError:
                self.wb = Workbook()
                self.wb.remove(self.wb['Sheet'])
        if len(data) <= 1:
            logging.warning('数据为空，未写入')
            return False

        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.create_sheet(sheet_name)
        ws.freeze_panes = 'A2'
        if ws.max_row != 1:
            data = data[1:]
        for row in data:
            ws.append(row)
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(data[0]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        for cell in ws[1]:
            if cell.value == '是否作弊':
                column_letter = cell.column_letter
                column_cells = ws[column_letter]
                for cell in column_cells[1:]:
                    if cell.value == '是':
                        cell.fill = styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        return True

    def save(self) -> bool:
        '''保存表格数据到本地'''     
        try:
            self.wb.active = 0
            self.wb.save(self.file_path)
            self.wb.close()
            return True
        except PermissionError as e:
            logging.error(f'文件保存失败!请勿在打开表格时操作：{e}')
        return False


class Schedule:
    def __init__(self, crontab: str, func: callable, *args, **kwargs) -> None:
        '''计划调用类'''
        self.crontab_expr = crontab
        self.exec = func
        if len(self.crontab_expr.split()) != 5:
            logging.warning('未正确设置schedule，故未启动计划')
            return
        self.cron = self.parse_crontab(self.crontab_expr)
        self.thread = threading.Thread(
            target=self.run,
            args=args,
            kwargs=kwargs,
            daemon=True,
        )
        logging.info(f'启动计划 [{self.crontab_expr}]')
        self.thread.start()

    def run(self, *args, **kwargs) -> None:
        '''执行函数'''
        while time.localtime().tm_sec != 0:
            time.sleep(1)
        while True:
            try:
                now = time.localtime()
                if (now.tm_min in self.cron[0] and
                        now.tm_hour in self.cron[1] and
                        now.tm_mday in self.cron[2] and
                        now.tm_mon in self.cron[3] and
                        now.tm_wday in self.cron[4]):
                    now_str = time.strftime('%Y-%m-%d %H:%M', now)
                    logging.info(f'执行计划[{self.crontab_expr}]')
                    threading.Thread(
                        target=self.exec,
                        args=args,
                        kwargs=kwargs,
                        daemon=True,
                    ).start()
                time.sleep(60)
            except:
                traceback.print_exc()

    def parse_crontab(self, crontab_expr: str) -> list:
        '''解析crontab'''
        fields = crontab_expr.split(' ')
        minute = self.parse_field(fields[0], 0, 59)
        hour = self.parse_field(fields[1], 0, 23)
        day_of_month = self.parse_field(fields[2], 1, 31)
        month = self.parse_field(fields[3], 1, 12)
        day_of_week = self.parse_field(fields[4], 0, 6)
        return (minute, hour, day_of_month, month, day_of_week)

    def parse_field(self, field: str, min_value: int, max_value: int):
        '''解析field'''
        if field == '*':
            return set(range(min_value, max_value + 1))
        values = set()
        for part in field.split(','):
            if '-' in part:
                start, end = map(int, part.split('-'))
                values.update(range(start, end + 1))
            else:
                values.add(int(part))
        return values


def recordInfo(bcz: BCZ, sqlite: SQLite):
    '''记录用户信息'''
    group_info_list = []
    for group_info in sqlite.queryObserveGroupInfo():
        group_info_list.append(bcz.getGroupInfo(group_info['shareKey']))
        logging.info(f'正在记录小班[{group_info["name"]}]的数据')
    sqlite.saveGroupInfo(group_info_list)


if __name__ == '__main__':
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.DEBUG)
    config = Config()
    recordInfo(BCZ(config), SQLite(config))